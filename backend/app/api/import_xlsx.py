from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import get_db
from app.auth import get_current_admin
from app.models import Equipment, EquipmentStatus, Warehouse, User
from openpyxl import load_workbook
import io

router = APIRouter(prefix="/api/import", tags=["import"])

REQUIRED_HEADERS = [
    "Штрихкод", "Наименование", "Категория", "Серийный номер",
    "Инвентарный номер", "Описание", "Текущий статус", "Склад"
]

STATUS_MAP = {
    "Рабочий": EquipmentStatus.WORKING,
    "Требует ремонта": EquipmentStatus.NEEDS_REPAIR,
    "В ремонте": EquipmentStatus.IN_REPAIR,
    "На складе": EquipmentStatus.IN_WAREHOUSE,
    "Выдан": EquipmentStatus.ISSUED,
    "Списан": EquipmentStatus.DECOMMISSIONED,
}

@router.post("/xlsx")
async def import_from_xlsx(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(400, "Файл должен быть в формате .xlsx")
    
    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active

    # Проверяем заголовки
    headers = [cell.value for cell in ws[1]]
    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        raise HTTPException(400, f"Отсутствуют обязательные столбцы: {', '.join(missing)}")

    # Собираем индексы колонок
    col_map = {h: headers.index(h) for h in REQUIRED_HEADERS}

    imported = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            barcode = str(row[col_map["Штрихкод"]]).strip() if row[col_map["Штрихкод"]] else None
            if not barcode:
                errors.append(f"Строка {row_idx}: пустой штрихкод")
                skipped += 1
                continue

            # Проверяем, существует ли уже оборудование с таким штрихкодом
            from sqlalchemy import select
            existing = await db.execute(select(Equipment).where(Equipment.barcode == barcode))
            if existing.scalar_one_or_none():
                errors.append(f"Строка {row_idx}: оборудование со штрихкодом {barcode} уже существует")
                skipped += 1
                continue

            # Ищем склад по названию
            warehouse_name = str(row[col_map["Склад"]]).strip() if row[col_map["Склад"]] else None
            warehouse_id = None
            if warehouse_name:
                wh = await db.execute(select(Warehouse).where(Warehouse.name == warehouse_name))
                wh_obj = wh.scalar_one_or_none()
                if wh_obj:
                    warehouse_id = wh_obj.id
                else:
                    errors.append(f"Строка {row_idx}: склад '{warehouse_name}' не найден, поле оставлено пустым")

            # Определяем статус
            status_str = str(row[col_map["Текущий статус"]]).strip() if row[col_map["Текущий статус"]] else "На складе"
            status = STATUS_MAP.get(status_str, EquipmentStatus.IN_WAREHOUSE)

            equipment = Equipment(
                barcode=barcode,
                name=str(row[col_map["Наименование"]]).strip() if row[col_map["Наименование"]] else "",
                category=str(row[col_map["Категория"]]).strip() if row[col_map["Категория"]] else None,
                serial_number=str(row[col_map["Серийный номер"]]).strip() if row[col_map["Серийный номер"]] else None,
                inventory_number=str(row[col_map["Инвентарный номер"]]).strip() if row[col_map["Инвентарный номер"]] else None,
                description=str(row[col_map["Описание"]]).strip() if row[col_map["Описание"]] else None,
                current_status=status,
                current_warehouse_id=warehouse_id
            )
            db.add(equipment)
            imported += 1
        except Exception as e:
            errors.append(f"Строка {row_idx}: {str(e)}")
            skipped += 1

    await db.commit()

    return {
        "message": "Импорт завершён",
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:20]  # Возвращаем первые 20 ошибок
    }
